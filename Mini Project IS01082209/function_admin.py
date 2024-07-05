import openpyxl
from tkinter import messagebox
import customtkinter as ctkt
from music import *

#-------------------------------------------------------------------------------------------
def view_customer_admin():
    def load_customers():
        try:
            workbook = openpyxl.load_workbook('customer.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            messagebox.showerror("Error", "No customer data found")
            view_customer_window.destroy()
            return
        
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            for j, value in enumerate(row):
                label = ctkt.CTkLabel(master=table_frame, text=value)
                label.grid(row=i, column=j, padx=5, pady=5)

    view_customer_window = ctkt.CTk()
    view_customer_window.geometry("700x450")
    view_customer_window.title("View Customer Information")

    table_frame = ctkt.CTkFrame(master=view_customer_window)
    table_frame.pack(pady=20)

    ctkt.CTkButton(view_customer_window, text="Load Customers", command=(lambda: (load_customers(), play_click()))).pack(pady=20)
    ctkt.CTkButton(view_customer_window, text="Back", command=(lambda: (view_customer_window.destroy(), play_click()))).pack(pady=10)

    view_customer_window.mainloop()


#-------------------------------------------------------------------------------------------
def edit_customer_admin():
    def load_customer_ids():
        try:
            workbook = openpyxl.load_workbook('customer.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            play_error()
            messagebox.showerror("Error", "Customer data not found")
            edit_customer_window.destroy()
            return

        customer_ids = [str(row[0]) for row in sheet.iter_rows(values_only=True) if row[0] is not None]
        customer_id_combobox.configure(values=customer_ids)

    def load_customer_information(event):
        selected_customer_id = customer_id_combobox.get()

        try:
            workbook = openpyxl.load_workbook('customer.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            play_error()
            messagebox.showerror("Error", "Customer data not found")
            edit_customer_window.destroy()
            return

        for row in sheet.iter_rows(values_only=True):
            if str(row[0]) == selected_customer_id:
                name_entry.delete(0, "end")
                phone_number_entry.delete(0, "end")
                age_entry.delete(0, "end")
                identification_number_entry.delete(0, "end")

                if row[1] is not None:
                    name_entry.insert(0, row[1])
                if row[2] is not None:
                    phone_number_entry.insert(0, row[2])
                if row[3] is not None:
                    age_entry.insert(0, row[3])
                if row[4] is not None:
                    identification_number_entry.insert(0, row[4])
                return

    def update_customer_information():
        customer_id = customer_id_combobox.get()
        name = name_entry.get()
        phone_number = phone_number_entry.get()
        age = age_entry.get()
        identification_number = identification_number_entry.get()

        if not customer_id:
            play_error()
            messagebox.showerror("Error", "Customer ID is required")
            return

        try:
            workbook = openpyxl.load_workbook('customer.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            play_error()
            messagebox.showerror("Error", "Customer data not found")
            edit_customer_window.destroy()
            return

        for row in sheet.iter_rows():
            if str(row[0].value) == customer_id:
                if name:
                    row[1].value = name
                if phone_number:
                    row[2].value = phone_number
                if age:
                    row[3].value = age
                if identification_number:
                    row[4].value = identification_number
                workbook.save('customer.xlsx')
                play_success()
                messagebox.showinfo("Success", "Customer information updated successfully")
                edit_customer_window.destroy()
                return
        play_error()
        messagebox.showerror("Error", f"Customer with ID {customer_id} not found")

    edit_customer_window = ctkt.CTk()
    edit_customer_window.geometry("800x500")
    edit_customer_window.title("Edit Customer Table")

    ctkt.CTkLabel(edit_customer_window, text="Select Customer ID:").pack(pady=5)
    customer_id_combobox = ctkt.CTkComboBox(edit_customer_window, state="readonly")
    customer_id_combobox.pack(pady=5)
    customer_id_combobox.bind("<<ComboboxSelected>>", load_customer_information)

    ctkt.CTkLabel(edit_customer_window, text="Name:").pack(pady=5)
    name_entry = ctkt.CTkEntry(edit_customer_window)
    name_entry.pack(pady=5)

    ctkt.CTkLabel(edit_customer_window, text="Phone Number:").pack(pady=5)
    phone_number_entry = ctkt.CTkEntry(edit_customer_window)
    phone_number_entry.pack(pady=5)

    ctkt.CTkLabel(edit_customer_window, text="Age:").pack(pady=5)
    age_entry = ctkt.CTkEntry(edit_customer_window)
    age_entry.pack(pady=5)

    ctkt.CTkLabel(edit_customer_window, text="Identification Number:").pack(pady=5)
    identification_number_entry = ctkt.CTkEntry(edit_customer_window)
    identification_number_entry.pack(pady=5)

    ctkt.CTkButton(edit_customer_window, text="Update", command=(lambda: (update_customer_information(), play_click()))).pack(pady=20)

    load_customer_ids()
    edit_customer_window.mainloop()

#-------------------------------------------------------------------------------------------
def view_room_admin():
    def load_room_status():
        try:
            workbook = openpyxl.load_workbook('room.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            messagebox.showerror("Error", "No room status data found")
            view_room_window.destroy()
            return
        
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            for j, value in enumerate(row):
                label = ctkt.CTkLabel(master=table_frame, text=value)
                label.grid(row=i, column=j, padx=5, pady=5)

    view_room_window = ctkt.CTk()
    view_room_window.geometry("1280x720")
    view_room_window.title("View Room Status")

    table_frame = ctkt.CTkFrame(master=view_room_window)
    table_frame.pack(pady=20)

    ctkt.CTkButton(view_room_window, text="Load Room Status", command=(lambda: (load_room_status(), play_click()))).pack(pady=20)
    ctkt.CTkButton(view_room_window, text="Back", command=(lambda: (view_room_window.destroy(), play_click()))).pack(pady=10)

    view_room_window.mainloop()

#-------------------------------------------------------------------------------------------
def edit_room_admin():
    def load_room_numbers():
        try:
            workbook = openpyxl.load_workbook('room.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            play_error()
            messagebox.showerror("Error", "Room data not found")
            edit_room_window.destroy()
            return

        room_numbers = [str(row[0].value) for row in sheet.iter_rows(min_row=2, values_only=False) if row[0].value is not None]
        room_number_combobox.configure(values=room_numbers)

    def load_customer_ids():
        try:
            workbook = openpyxl.load_workbook('customer.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            play_error()
            messagebox.showerror("Error", "Customer data not found")
            edit_room_window.destroy()
            return

        customer_ids = [str(row[0].value) for row in sheet.iter_rows(min_row=2, values_only=False) if row[0].value is not None]
        customer_id_combobox.configure(values=customer_ids)

    def update_room_status():
        room_number = room_number_combobox.get()
        status = status_combobox.get()
        customer_id = customer_id_combobox.get()

        if not room_number or not status or (not customer_id and status.lower() == "occupied"):
            play_error()
            messagebox.showerror("Error", "All fields are required")
            return

        try:
            workbook = openpyxl.load_workbook('room.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            play_error()
            messagebox.showerror("Error", "Room data not found")
            edit_room_window.destroy()
            return

        for row in sheet.iter_rows(min_row=2, values_only=False):
            if str(row[0].value) == room_number:
                row[3].value = status 
                row[4].value = customer_id 
                workbook.save('room.xlsx')
                play_success()
                messagebox.showinfo("Success", "Room status updated successfully")
                edit_room_window.destroy()
                return
        play_error()
        messagebox.showerror("Error", f"Room with number {room_number} not found")

    edit_room_window = ctkt.CTk()
    edit_room_window.geometry("800x500")
    edit_room_window.title("Edit Room Table")

    ctkt.CTkLabel(edit_room_window, text="Room Number:").pack(pady=5)
    room_number_combobox = ctkt.CTkComboBox(edit_room_window, state="readonly")
    room_number_combobox.pack(pady=5)

    ctkt.CTkLabel(edit_room_window, text="Status:").pack(pady=5)
    status_combobox = ctkt.CTkComboBox(edit_room_window, values=["Unavailable", "Available", "Occupied", "Maintenance"], state="readonly")
    status_combobox.pack(pady=5)

    ctkt.CTkLabel(edit_room_window, text="Customer ID (if occupied):").pack(pady=5)
    customer_id_combobox = ctkt.CTkComboBox(edit_room_window, state="readonly")
    customer_id_combobox.pack(pady=5)

    ctkt.CTkButton(edit_room_window, text="Update", command=lambda: (update_room_status(), play_click())).pack(pady=20)

    load_room_numbers()
    load_customer_ids()
    edit_room_window.mainloop()

#-------------------------------------------------------------------------------------------

def delete_customer_admin():
    def load_customers():
        try:
            workbook = openpyxl.load_workbook('customer.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            play_error()
            messagebox.showerror("Error", "No customer data found")
            delete_customer_window.destroy()
            return
      
        header_labels = [cell.value for cell in sheet[1]]

        for col, header in enumerate(header_labels):
            label = ctkt.CTkLabel(master=table_frame, text=header, font=('Arial', 10, 'bold'))
            label.grid(row=0, column=col, padx=5, pady=5)

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            for j, value in enumerate(row):
                label = ctkt.CTkLabel(master=table_frame, text=value)
                label.grid(row=i, column=j, padx=5, pady=5)

    def delete_customer():
        customer_id = entry_customer_id.get()

        if not customer_id:
            play_error()
            messagebox.showerror("Error", "Please enter a customer ID")
            return

        try:
            workbook = openpyxl.load_workbook('customer.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            play_error()
            messagebox.showerror("Error", "Customer data not found")
            delete_customer_window.destroy()
            return

        customer_found = False
        for row in sheet.iter_rows():
            if str(row[0].value) == customer_id:
                sheet.delete_rows(row[0].row)
                workbook.save('customer.xlsx')
                customer_found = True
                break

        if customer_found:
            play_success()
            renumber_customer_ids()
            messagebox.showinfo("Success", "Customer deleted successfully")
            delete_customer_window.destroy()
        else:
            play_error()
            messagebox.showerror("Error", f"Customer with ID {customer_id} not found")

    def renumber_customer_ids():
        try:
            workbook = openpyxl.load_workbook('customer.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            play_error()
            messagebox.showerror("Error", "Customer data not found")
            return

        for i, row in enumerate(sheet.iter_rows(min_row=2), start=1):
            row[0].value = str(i)

        workbook.save('customer.xlsx')

    delete_customer_window = ctkt.CTk()
    delete_customer_window.geometry("700x450")
    delete_customer_window.title("Delete Customer")

    table_frame = ctkt.CTkFrame(master=delete_customer_window)
    table_frame.pack(pady=20)

    load_customers()

    ctkt.CTkLabel(delete_customer_window, text="Enter Customer ID to delete:").pack(pady=5)
    entry_customer_id = ctkt.CTkEntry(delete_customer_window)
    entry_customer_id.pack(pady=5)

    ctkt.CTkButton(delete_customer_window, text="Delete Customer", command=(lambda: (delete_customer(), play_click()))).pack(pady=10)

    delete_customer_window.mainloop()

#-------------------------------------------------------------------------------------------
def generate_reports():
    try:
        workbook_customer = openpyxl.load_workbook('customer.xlsx')
        workbook_room = openpyxl.load_workbook('room.xlsx')
        sheet_customer = workbook_customer.active
        sheet_room = workbook_room.active
    except FileNotFoundError:
        play_error()
        messagebox.showerror("Error", "Data not found")
        return

    total_customers = sheet_customer.max_row - 1

    total_rooms = sheet_room.max_row - 1

    room_types = set(row[2] for row in sheet_room.iter_rows(min_row=2, max_col=3, values_only=True))
    total_room_types = len(room_types)

    room_count_by_type = {room_type: 0 for room_type in room_types}
    for row in sheet_room.iter_rows(min_row=2, max_col=3, values_only=True):
        room_type = row[2]
        room_count_by_type[room_type] += 1

    report = f"Customer Summary:\n\nTotal Customers: {total_customers}\n\n"
    report += f"Room Summary:\n\nTotal Rooms: {total_rooms}\nTotal Room Types: {total_room_types}\n\n"

    report += "Total Rooms by Room Type:\n"
    for room_type, count in room_count_by_type.items():
        report += f"{room_type}: {count}\n"

    play_success()
    messagebox.showinfo("Reports", report)
