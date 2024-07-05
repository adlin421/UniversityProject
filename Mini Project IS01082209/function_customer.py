import tkinter as tk
from tkinter import messagebox
import customtkinter as ctk
from openpyxl import load_workbook
import openpyxl
from music import *

def display_facility_schedule():
    def load_facility_schedule():
        try:
            workbook = openpyxl.load_workbook('facility_schedule.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            messagebox.showerror("Error", "No facility schedule data found")
            play_error()
            view_schedule_window.destroy()
            return
        
      
        for widget in table_frame.winfo_children():
            widget.destroy()
        
     
        headers = ('Type of Sport', 'Time', 'Venue', 'Staff Incharge')
        for j, header in enumerate(headers):
            header_label = ctk.CTkLabel(master=table_frame, text=header)
            header_label.grid(row=0, column=j, padx=5, pady=5)

      
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            for j, value in enumerate(row):
                label = ctk.CTkLabel(master=table_frame, text=value)
                label.grid(row=i, column=j, padx=5, pady=5)

    view_schedule_window = ctk.CTk()
    view_schedule_window.geometry("1280x720")
    view_schedule_window.title("View Facility Schedule")

    table_frame = ctk.CTkFrame(master=view_schedule_window)
    table_frame.pack(pady=20)

    ctk.CTkButton(view_schedule_window, text="Load Facility Schedule", command=(lambda: (load_facility_schedule(), play_click()))).pack(pady=20)
    ctk.CTkButton(view_schedule_window, text="Back", command=(lambda: (view_schedule_window.destroy(), play_click()))).pack(pady=10)

    view_schedule_window.mainloop()



def edit_customer_status():
    def save_status():
        phone_number = phone_entry.get()
        new_status = status_var.get()

        workbook = load_workbook('customer.xlsx')
        sheet = workbook.active

        phone_column_index = 3  
        status_column_index = 6 

        found = False
        for row in range(2, sheet.max_row + 1):  
            cell_value = sheet.cell(row=row, column=phone_column_index).value
            if cell_value == phone_number:
                sheet.cell(row=row, column=status_column_index, value=new_status)
                play_success()
                workbook.save('customer.xlsx')
                messagebox.showinfo("Success", "Status updated successfully!")
                found = True
                break

        if not found:
            play_error()
            messagebox.showerror("Error", "Phone number not found.")
        
        edit_window.destroy()

    edit_window = ctk.CTk()
    edit_window.geometry("400x300")
    edit_window.title("Edit Customer Status")

    ctk.CTkLabel(edit_window, text="Edit Status", font=("Arial", 16)).pack(pady=20)

    ctk.CTkLabel(edit_window, text="Phone Number:").pack(pady=5)
    phone_entry = ctk.CTkEntry(edit_window)
    phone_entry.pack(pady=5)

    ctk.CTkLabel(edit_window, text="New Status:").pack(pady=5)
    status_var = tk.StringVar(value="None")
    status_options = ["Checked In", "On Hold","Checked Out"]  
    status_menu = ctk.CTkOptionMenu(edit_window, variable=status_var, values=status_options,command=play_click())
    status_menu.pack(pady=5)

    save_button = ctk.CTkButton(edit_window, text="Save", command=(lambda: (save_status(), play_click())))
    save_button.pack(pady=20)
    ctk.CTkButton(edit_window, text="Back", command=(lambda: (edit_window.destroy(), play_click()))).pack(pady=20)

    edit_window.mainloop()