import openpyxl
from tkinter import messagebox
import customtkinter as ctkt
from tkinter import StringVar
from tkinter import ttk
from music import play_click, play_error, play_success

def add_customer_management():
    def get_next_customer_id():
        try:
            workbook = openpyxl.load_workbook('customer.xlsx')
            sheet = workbook.active
            return str(sheet.max_row)
        except FileNotFoundError:
            return "1"

    def submit_customer():
        customer_id = get_next_customer_id()
        name = name_entry.get()
        phone_number = phone_number_entry.get()
        age = age_entry.get()
        identification_number = identification_number_entry.get()

        if not name or not phone_number or not age or not identification_number:
            play_error()
            messagebox.showerror("Error", "All fields are required")
            return

        if not name.isalpha() or len(name) > 20:
            play_error()
            messagebox.showerror("Error", "Name must be alphabetic and maximum 20 characters long")
            return

        if not phone_number.isdigit() or len(phone_number) > 12:
            play_error()
            messagebox.showerror("Error", "Phone number must be less than 12 digits long")
            return

        if not age.isdigit() or len(age) > 3:
            play_error()
            messagebox.showerror("Error", "Age must be a number and maximum 3 digits long")
            return

        if not identification_number.isdigit() or len(identification_number) > 13:
            play_error()
            messagebox.showerror("Error", "Identification number must be less than 13 digits long")
            return

        try:
            workbook = openpyxl.load_workbook('customer.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            play_error()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Customer ID", "Name", "Phone Number", "Age", "Identification Number"])

        sheet.append([customer_id, name, phone_number, age, identification_number])
        workbook.save('customer.xlsx')
        play_success()
        messagebox.showinfo("Success", "Customer information added successfully")
        add_customer_window.destroy()

    add_customer_window = ctkt.CTk()
    add_customer_window.geometry("700x500")
    add_customer_window.title("Add Customer Information")

    ctkt.CTkLabel(add_customer_window, text="Name:").pack(pady=5)
    name_entry = ctkt.CTkEntry(add_customer_window)
    name_entry.pack(pady=5)

    ctkt.CTkLabel(add_customer_window, text="Phone Number:").pack(pady=5)
    phone_number_entry = ctkt.CTkEntry(add_customer_window)
    phone_number_entry.pack(pady=5)

    ctkt.CTkLabel(add_customer_window, text="Age:").pack(pady=5)
    age_entry = ctkt.CTkEntry(add_customer_window)
    age_entry.pack(pady=5)

    ctkt.CTkLabel(add_customer_window, text="Identification Number:").pack(pady=5)
    identification_number_entry = ctkt.CTkEntry(add_customer_window)
    identification_number_entry.pack(pady=5)

    ctkt.CTkButton(add_customer_window, text="Submit", command=(lambda: (submit_customer(), play_click()))).pack(pady=20)

    add_customer_window.mainloop()

def view_customer_management():
    def load_customers():
        try:
            workbook = openpyxl.load_workbook('customer.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            play_error()
            messagebox.showerror("Error", "No customer data found")
            view_customer_window.destroy()
            return

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            for j, value in enumerate(row):
                label = ctkt.CTkLabel(master=table_frame, text=value)
                label.grid(row=i, column=j, padx=5, pady=5)

    view_customer_window = ctkt.CTk()
    view_customer_window.geometry("700x600")
    view_customer_window.title("View Customer Information")

    table_frame = ctkt.CTkFrame(master=view_customer_window)
    table_frame.pack(pady=20)

    ctkt.CTkButton(view_customer_window, text="Load Customers", command=(lambda: (load_customers(), play_click()))).pack(pady=20)
    ctkt.CTkButton(view_customer_window, text="Back", command=(lambda: (view_customer_window.destroy(), play_click()))).pack(pady=10)

    view_customer_window.mainloop()

def edit_room_management():
    def load_room_numbers():
        try:
            workbook = openpyxl.load_workbook('room.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            play_error()
            messagebox.showerror("Error", "Room data not found")
            edit_room_window.destroy()
            return

        room_numbers = [str(row[0].value) for row in sheet.iter_rows(min_row=2, values_only=False) if row[0].value is not None]
        room_number_combobox.configure(values=room_numbers)

    def load_customer_ids():
        try:
            workbook = openpyxl.load_workbook('customer.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            play_error()
            messagebox.showerror("Error", "Customer data not found")
            edit_room_window.destroy()
            return

        customer_ids = [str(row[0].value) for row in sheet.iter_rows(min_row=2, values_only=False) if row[0].value is not None]
        customer_id_combobox.configure(values=customer_ids)

    def update_room_status():
        room_number = room_number_combobox.get()
        status = status_combobox.get()
        customer_id = customer_id_combobox.get()

        if not room_number or not status or (not customer_id and status.lower() == "occupied"):
            play_error()
            messagebox.showerror("Error", "All fields are required")
            return

        try:
            workbook = openpyxl.load_workbook('room.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            play_error()
            messagebox.showerror("Error", "Room data not found")
            edit_room_window.destroy()
            return

        for row in sheet.iter_rows(min_row=2, values_only=False):
            if str(row[0].value) == room_number:
                row[3].value = status 
                row[4].value = customer_id 
                workbook.save('room.xlsx')
                play_success()
                messagebox.showinfo("Success", "Room status updated successfully")
                edit_room_window.destroy()
                return
        play_error()
        messagebox.showerror("Error", f"Room with number {room_number} not found")

    edit_room_window = ctkt.CTk()
    edit_room_window.geometry("800x500")
    edit_room_window.title("Edit Room Table")

    ctkt.CTkLabel(edit_room_window, text="Room Number:").pack(pady=5)
    room_number_combobox = ctkt.CTkComboBox(edit_room_window, state="readonly")
    room_number_combobox.pack(pady=5)

    ctkt.CTkLabel(edit_room_window, text="Status:").pack(pady=5)
    status_combobox = ctkt.CTkComboBox(edit_room_window, values=["Unavailable", "Available", "Occupied", "Maintenance"], state="readonly")
    status_combobox.pack(pady=5)

    ctkt.CTkLabel(edit_room_window, text="Customer ID (if occupied):").pack(pady=5)
    customer_id_combobox = ctkt.CTkComboBox(edit_room_window, state="readonly")
    customer_id_combobox.pack(pady=5)

    ctkt.CTkButton(edit_room_window, text="Update", command=lambda: (update_room_status(), play_click())).pack(pady=20)

    load_room_numbers()
    load_customer_ids()
    edit_room_window.mainloop()



def view_room_management():
    def load_room_status():
        try:
            workbook = openpyxl.load_workbook('room.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            play_error()
            messagebox.showerror("Error", "No room status data found")
            view_room_window.destroy()
            return

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            for j, value in enumerate(row):
                label = ctkt.CTkLabel(master=table_frame, text=value)
                label.grid(row=i, column=j, padx=5, pady=5)

    view_room_window = ctkt.CTk()
    view_room_window.geometry("700x1000")
    view_room_window.title("View Room Status")

    table_frame = ctkt.CTkFrame(master=view_room_window)
    table_frame.pack(pady=20)

    ctkt.CTkButton(view_room_window, text="Load Room Status", command=(lambda: (load_room_status(), play_click()))).pack(pady=20)
    ctkt.CTkButton(view_room_window, text="Back", command=(lambda: (view_room_window.destroy(), play_click()))).pack(pady=10)

    view_room_window.mainloop()
